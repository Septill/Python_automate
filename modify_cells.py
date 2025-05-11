from openpyxl import Workbook

wb = Workbook()
ws = wb.active

x = 1
for i in range (1,11):
    for j in range(1,11):
        ws.cell(i,j,x)
        x += 1

ws.merge_cells("b2:d4")
#ws.unmerge_cells("b2:d4")

ws.insert_cols(2,3)
ws.insert_rows(4,1)

ws.delete_cols(2,3)
ws.delete_rows(4,1)

wb.save(r"D:\Python workspace\test python\test_cell.xlsx")