from openpyxl import Workbook

wb = Workbook()

ws1 = wb.active
print(ws1.title)

ws2 = wb.create_sheet("file2",1)
ws3 = wb.create_sheet("file3", 2)

wb.move_sheet("file2", -1)

#del wb["file3"]

print(wb.sheetnames)