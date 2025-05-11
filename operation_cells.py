from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["c3"] = "Li LIANG"

cell = ws.cell(5,3,"LIANG Li")

cell2= ws.cell(6,4)
cell2.value = "new cell"

print(cell.coordinate) #coordinate for this cell
print(cell.row)
print(cell.column)

x = 1
for i in range (6,15):
    for j in range (7,11):
        ws.cell(i,j,x)
        x += 1

print(ws["a6:c8"])

wb.save(r"D:\Python workspace\test python\test.xlsx")