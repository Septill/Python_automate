from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

wb = Workbook ()
ws = wb.active

ws["a1"] = "LIANG"

ws["b2"] = "Li"

f = Font(size=15,color = "0000FF", bold= True)

ws["a1"].font = f
ws.row_dimensions [2].height =  30
ws.column_dimensions ["D"] .width = 30

a = Alignment(horizontal="left", vertical="top", text_rotation= 90, wrap_text= True)

ws["b2"].alignment = a

wb.save(r"D:\Python workspace\test python\test_font.xlsx")