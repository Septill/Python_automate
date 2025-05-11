from openpyxl import Workbook
from openpyxl.utils import FORMULAE
from openpyxl.formula.translate import Translator

wb = Workbook()
ws = wb.active

print(len(FORMULAE))
print("SUM" in FORMULAE)

ws.append(["num1","num2","sum","avrage"])
ws.append([56,78])
ws.append([25,3])
ws.append([82,68])

ws["c2"] = "=SUM(A2+B2)"
ws["d2"] = "=AVERAGE(A2:B2)"

#ws["c3"] = Translator(formula="=SUM(A2+B2)", origin="c2").translate_formula("c3")
#ws["c4"] = Translator(formula="=SUM(A2+B2)", origin="c2").translate_formula("c4")

#ws["d3"] = Translator(formula="=AVERAGE(A2:B2)", origin="d2").translate_formula("d3")
#ws["d4"] = Translator(formula="=AVERAGE(A2:B2)", origin="d2").translate_formula("d4")

for cell in ws["c3:c4"] :
    cell[0].value = Translator(formula="=SUM(A2+B2)", origin="c2").translate_formula(cell[0].coordinate)

for cell in ws["d3:d4"] :
    cell[0].value = Translator(formula="=AVERAGE(A2:B2)", origin="d2").translate_formula(cell[0].coordinate)

Column_cells = ws["C"]

print(Column_cells)

wb.save(r"D:\Python workspace\test python\test_formule.xlsx")